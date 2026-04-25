from importlib.resources import Package
from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from .models import LoanHead 
from .models import LoanCategory 
from .models import LoanPackage
from .forms import LoanHeadForm, LoanCategoryForm, LoanPackageForm
from openpyxl import Workbook
from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login
from django.contrib import messages
from django.contrib.auth import logout
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required
from smart_finance.models import ContactMessage
from admin_panel.models import LoanHead
from django.db.models import Count
from django.template.loader import get_template
from xhtml2pdf import pisa
from io import BytesIO


def dashboard(request):

    if not request.session.get("admin_logged_in"):
        return redirect("admin_panel:login")

    context = {
        "heads_count": LoanHead.objects.count(),
        "categories_count": LoanCategory.objects.count(),
        "packages_count": LoanPackage.objects.count(),
    }

    return render(request, "admin_panel/dashboard.html", context)

# ---------------- LOAN HEAD -----------------

def loan_head_list(request):
    heads = LoanHead.objects.all()
    return render(request, "admin_panel/loan_head_list.html", {"heads": heads})


def loan_head_create(request):
    form = LoanHeadForm(request.POST or None)

    if form.is_valid():
        form.save()
        return redirect("admin_panel:loan_head_list")

    return render(request, "admin_panel/loan_head_form.html", {"form": form})


def loan_head_edit(request, pk):
    head = get_object_or_404(LoanHead, pk=pk)
    form = LoanHeadForm(request.POST or None, instance=head)

    if form.is_valid():
        form.save()
        return redirect("admin_panel:loan_head_list")

    return render(request, "admin_panel/loan_head_form.html", {"form": form})


def loan_head_delete(request, pk):
    head = get_object_or_404(LoanHead, pk=pk)
    head.delete()
    return redirect("admin_panel:loan_head_list")

def loan_head_report(request):
    heads = LoanHead.objects.all()

    return render(request, "admin_panel/loan_head_report.html", {
        "heads": heads
    })

def loan_head_export_excel(request):

    wb = Workbook()
    ws = wb.active
    ws.title = "Loan Heads"

    # Header row
    ws.append(["#", "Loan Name"])

    # Data
    heads = LoanHead.objects.all()

    for idx, head in enumerate(heads, start=1):
        ws.append([idx, head.name])

    # Response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="Loan_Heads_Report.xlsx"'

    wb.save(response)
    return response


# ---------------- CATEGORY -----------------

def category_list(request):
    categories = LoanCategory.objects.all()
    return render(request, "admin_panel/category_list.html", {"categories": categories})


def category_create(request):
    form = LoanCategoryForm(request.POST or None)

    if form.is_valid():
        form.save()
        return redirect("admin_panel:category_list")

    return render(
        request,
        "admin_panel/category_form.html",
        {"form": form}
    )

def category_edit(request, pk):
    category = get_object_or_404(LoanCategory, pk=pk)

    form = LoanCategoryForm(request.POST or None, instance=category)

    if form.is_valid():
        form.save()
        return redirect("admin_panel:category_list")

    return render(request, "admin_panel/category_form.html", {
        "form": form
    })


def category_delete(request, pk):
    category = get_object_or_404(LoanCategory, pk=pk)
    category.delete()
    return redirect("admin_panel:category_list")

def category_export_excel(request):

    wb = Workbook()
    ws = wb.active
    ws.title = "Loan Categories"

    # Header row
    ws.append(["#", "Category Name", "Description"])

    categories = LoanCategory.objects.all()

    for idx, cat in enumerate(categories, start=1):
        ws.append([
            idx,
            cat.name,
            cat.description or ""
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response["Content-Disposition"] = (
        'attachment; filename="Loan_Categories_Report.xlsx"'
    )

    wb.save(response)
    return response



# ---------------- PACKAGE -----------------

def package_list(request):
    packages = LoanPackage.objects.select_related("category").all()

    return render(
        request,
        "admin_panel/package_list.html",
        {
            "packages": packages
        }
    )


def package_create(request):
    form = LoanPackageForm(request.POST or None)

    if form.is_valid():
        form.save()
        return redirect("admin_panel:package_list")

    return render(request, "admin_panel/package_form.html", {"form": form})


def package_edit(request, pk):
    package = get_object_or_404(LoanPackage, pk=pk)

    form = LoanPackageForm(request.POST or None, instance=package)

    if form.is_valid():
        form.save()
        return redirect("admin_panel:package_list")

    return render(
        request,
        "admin_panel/package_form.html",
        {
            "form": form,
            "edit_mode": True
        }
    )


def package_delete(request, pk):
    package = get_object_or_404(LoanPackage, pk=pk)

    package.delete()

    return redirect("admin_panel:package_list")

def package_export_excel(request):

    wb = Workbook()
    ws = wb.active
    ws.title = "Loan Packages"

    # Header row
    ws.append(["#", "Package Title", "Category"])

    packages = LoanPackage.objects.select_related("category").all()

    for idx, pkg in enumerate(packages, start=1):
        ws.append([
            idx,
            pkg.title,
            pkg.category.name if pkg.category else ""
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response["Content-Disposition"] = (
        'attachment; filename="Loan_Packages_Report.xlsx"'
    )

    wb.save(response)
    return response

def admin_login(request):

    if request.method == "POST":

        username = request.POST.get("email")
        password = request.POST.get("password")

        # Fixed admin credentials
        if username == "admin@gmail.com" and password == "admin123":

            request.session["admin_logged_in"] = True

            return redirect("admin_panel:dashboard")

        else:
            messages.error(request, "Invalid Admin Credentials")

    return render(request, "admin_panel/admin_login.html")

def admin_logout(request):
    request.session.flush()
    return redirect("admin_panel:login")




def contact_list(request):

    contacts = ContactMessage.objects.all().order_by("-created_at")

    return render(
        request,
        "admin_panel/contact_list.html",
        {"contacts": contacts}
    )


def contact_delete(request, pk):

    contact = get_object_or_404(ContactMessage, pk=pk)
    contact.delete()

    return redirect("admin_panel:contact_list")



def admin_report(request):

    search = request.GET.get("search")
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")
    category_id = request.GET.get("category")
    data_type = request.GET.get("type")  # NEW

    # ---------------- LOAN HEAD ----------------
    heads = LoanHead.objects.all()

    if search:
        heads = heads.filter(name__icontains=search)

    if start_date:
        heads = heads.filter(created_at__date__gte=start_date)

    if end_date:
        heads = heads.filter(created_at__date__lte=end_date)

    # ---------------- CATEGORY ----------------
    categories = LoanCategory.objects.all()

    if search:
        categories = categories.filter(name__icontains=search)

    # ---------------- PACKAGE ----------------
    packages = LoanPackage.objects.select_related("category").all()

    if search:
        packages = packages.filter(title__icontains=search)

    if category_id:
        packages = packages.filter(category_id=category_id)

    # ---------------- CONTACT ----------------
    contacts = ContactMessage.objects.all().order_by("-created_at")

    if start_date:
        contacts = contacts.filter(created_at__date__gte=start_date)

    if end_date:
        contacts = contacts.filter(created_at__date__lte=end_date)

    return render(request, "admin_panel/admin_report.html", {
        "heads": heads,
        "categories": categories,
        "packages": packages,
        "contacts": contacts,
        "all_categories": LoanCategory.objects.all(),
        "data_type": data_type
    })





def admin_report_pdf(request):

    # 🔍 GET FILTERS
    search = request.GET.get("search")
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")
    category_id = request.GET.get("category")

    # ---------------- LOAN HEAD ----------------
    heads = LoanHead.objects.all()

    if search:
        heads = heads.filter(name__icontains=search)

    if start_date:
        heads = heads.filter(created_at__date__gte=start_date)

    if end_date:
        heads = heads.filter(created_at__date__lte=end_date)

    # ---------------- CATEGORY ----------------
    categories = LoanCategory.objects.all()

    if search:
        categories = categories.filter(name__icontains=search)

    if category_id:
        categories = categories.filter(id=category_id)

    # ---------------- PACKAGE ----------------
    packages = LoanPackage.objects.select_related("category").all()

    if search:
        packages = packages.filter(title__icontains=search)

    if category_id:
        packages = packages.filter(category_id=category_id)

    # ---------------- CONTACT ----------------
    contacts = ContactMessage.objects.all().order_by("-created_at")

    if start_date:
        contacts = contacts.filter(created_at__date__gte=start_date)

    if end_date:
        contacts = contacts.filter(created_at__date__lte=end_date)

    # 🧾 RENDER TEMPLATE
    template = get_template("admin_panel/admin_report_pdf.html")

    html = template.render({
        "heads": heads,
        "categories": categories,
        "packages": packages,
        "contacts": contacts,
    })

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="admin_report.pdf"'

    pisa.CreatePDF(BytesIO(html.encode("UTF-8")), dest=response)

    return response